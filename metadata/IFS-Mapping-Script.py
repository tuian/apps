
import docx
import openpyxl #install openpyxl first 
import os
import xlsxwriter 
import sys  
import datetime
import re
import warnings
reload(sys)  
sys.setdefaultencoding('utf8')
from docx.document import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph
#parentDir = "..\\Desktop\\Metadata Repository" #change path of parent directory
#parentDir = "C:\IFS\\"
parentDir = "C:\IFS\Automation_After_Corrections\\"
input_excel_filename = "IFS Documents Analysis.xlsx"
input_excel_filepath =  parentDir + input_excel_filename

rownum = 1
rownum_IFS = 1
rownum_XSD = 1
rowmap=1
col = 0
addtext = ""

flag = 'false'
#flag = 'true'

def removeSpaces(input_str):

	# clean the input string with any space characters in the middle (e.g. 'rep/data/user_list/user/user_head_list/user_head/ job_auth_level_id/ annot/ctx/id')
	pattern = re.compile(r'\s+')
	input_str = re.sub(pattern, '', input_str)

	return input_str

def Print(*str):
    global flag
    if(flag == 'true'):
        print str

def processXSDAttribute(input_str):

	"""START"""
	print "\t\t Original Attribute Name: "+ input_str
		  #+ " Last Character = " + input_str[-1]

	# strip the '/' character if present in the begining (e.g. //rep/data/report/report_foot_list/)
	if (input_str[-1] == '/'):
		input_str = input_str[:-1]

	# clean the input string with any space characters in the middle (e.g. 'rep/data/user_list/user/user_head_list/user_head/ job_auth_level_id/ annot/ctx/id')
	pattern = re.compile(r'\s+')
	input_str = re.sub(pattern, '', input_str)

	#pattern 1: '/val'
	if(re.match(r'.*(/val)$', input_str) ):
		input_str = input_str[:-4]
	else:
    		# pattern 2: '/annot/ctx/id'
			if (re.match(r'.*(/annot/ctx/id)$', input_str)):
				input_str = input_str[:-13]

	output_str =  (input_str.split('/'))[-1]
	print "\t\t\t Extracted Attribute Name: " + output_str
	"""END"""

	return output_str

def iter_block_items(parent):
    """
    Yield each paragraph and table child within *parent*, in document order.
    Each returned value is an instance of either Table or Paragraph. *parent*
    would most commonly be a reference to a main Document object, but
    also works for a _Cell object, which itself can contain paragraphs and tables.
    """
    if isinstance(parent, Document):
        parent_elm = parent.element.body
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        raise ValueError("something's not right")

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)

def object_create(objectType):

#sheet variable refers to the input excel sheet
 	Print("Object Create Function  Invoked. ObjectType: ",objectType)
	#Print("Sheet Max Row =",sheet.max_row)
	#print "Sheet Min &  Max Row =",sheet.min_row,sheet.max_row

	#from row 2 to max row in the input excel sheet. Note: Row 1 has the header information, so start from row 2
	for row in range(2, (sheet.max_row+1)):

		#print "File Name: " + sheet['C' + str(row)].value

		# Check in Column C2,C3... for the Status Flag. If Open Process the document
		if ( sheet['C' + str(row)].value != None):

			if((sheet['B' + str(row)].value).lower() == "Open".lower()):



				#get name of ifs doc to read
				global parentDir
				docName =  parentDir+"\\"+ sheet['C' + str(row)].value

				#add .docx extension
				docName = docName + '.docx'

				print "\t Processing IFS Document: " + docName

				#read the document
				try:
					doc1 = docx.Document(docName)
				except:
					print "ERROR in reading a document. Check if the file exists or the filename"
					workbook.close()
					exit()
				#read the section name where we have the table
				#strip the whitespaces and convert to all upper case
				sectionName = (sheet['E' + str(row)].value.encode("utf-8").strip()).upper()
				Print("Section name given in the input sheet",sectionName)

				ifsName=sheet['D' + str(row)].value.encode("utf-8").strip()
				xsdName=sheet['I' + str(row)].value.encode("utf-8").strip()
				ifsDescription = sheet['F' + str(row)].value.encode("utf-8").strip()
				sharepoint_pathUrl= sheet['G' + str(row)].value

				avaloq_service_name = ""
				avaloq_service_name = sheet['H' + str(row)].value

				#object creation for ifs
		
				#initialise the flag to false
				#when the section name is matched set the flag to true
				ispara = 'false'

				global rownum
				global rownum_IFS
				global rownum_XSD
				global rowmap
		
		
				#for each block in the word document, check if the block is a paragraph or table
				#if table, process the table rows and columns


				for block in iter_block_items(doc1):
						#Check if the block is table
						if (isinstance(block, Table)):

							Print("Table found")

							#check if this is the table we wanted (i.e. section name matches the input sheet)
							if ispara == 'true' :

								Print("Table we wanted...")

								ifsTable= block
								count = 0
								entityList=[];

								for row in ifsTable.rows:

									Print("Count: ", count)
									Print(row.cells[0].text.encode("utf-8"), row.cells[1].text.encode("utf-8"), row.cells[2].text.encode("utf-8") , row.cells[3].text.encode("utf-8"), row.cells[4].text.encode("utf-8"),row.cells[5].text.encode("utf-8"))

									entityName= re.sub("[\(\[].*?[\)\]]", "", row.cells[0].text.encode("utf-8"))

									Print("Entity Name:",entityName, " Count: ",count)

									if (count>0):

										Print(row.cells[2].text.encode("utf-8"))

										if(row.cells[2].text.encode("utf-8") != "Aggregate") :

											if (objectType== 'IFS'):

												Print("Inside IFS")
												Print("Description:",row.cells[1].text.encode("utf-8"))
												Print("Cardinality:",row.cells[3].text.encode("utf-8"))

												if row.cells[1].text.encode("utf-8") != row.cells[3].text.encode("utf-8") :

													if row.cells[0].text.encode("utf-8") not in entityList :
														Print("Not in entityList",row.cells[0].text.encode("utf-8"))
														Print("Insert IFS Object: ", rownum_IFS, " Entity Name ", entityName,"Entity Type: Attribute")

														entityList.append(row.cells[0].text.encode("utf-8"))

														worksheet_IFS.write_string(rownum_IFS ,col + 0 , objectType.strip());
														worksheet_IFS.write_string(rownum_IFS ,col + 1 , 'Default');
														worksheet_IFS.write_string(rownum_IFS ,col + 2 ,ifsName.strip() );
														worksheet_IFS.write_string(rownum_IFS ,col + 3 , entityName.strip())
														worksheet_IFS.write_string(rownum_IFS ,col + 6 , 'Attribute');
														# description
														worksheet_IFS.write_string(rownum_IFS ,col + 7 ,  row.cells[1].text.encode("utf-8"));
														worksheet_IFS.write_string(rownum_IFS, col + 10, 'Phase 2');

														rownum_IFS = rownum_IFS + 1
													else:
														Print("Already in entityList",row.cells[0].text.encode("utf-8"))

													#create the mappings
													Print("Insert Mapping: ",rowmap," Entity Name ",entityName)

													worksheet2.write_string(rowmap ,col + 0 , 'IFS');
													worksheet2.write_string(rowmap ,col + 1 , 'Default');

													#ifs name from the input file
													worksheet2.write_string(rowmap ,col + 2 , ifsName.strip());

													# IFS (source) attribute name from the word document table
													worksheet2.write_string(rowmap ,col + 3 ,  entityName.strip());

													worksheet2.write_string(rowmap ,col + 4 , 'ABS Dev');
													worksheet2.write_string(rowmap ,col + 5 , 'XSD');

													#XSD Name from the input file
													worksheet2.write_string(rowmap ,col + 6 , xsdName.strip());

													# XSD (target) attribute name from the word document table. process the string to extract the desired substring
													#worksheet2.write_string(rowmap ,col + 7 ,entityName.strip());
													xsd_target_attribute = (row.cells[5].text.encode("utf-8")).strip()
													worksheet2.write_string(rowmap, col + 7, processXSDAttribute(xsd_target_attribute));

													#description
													worksheet2.write_string(rowmap ,col + 8 ,row.cells[1].text.encode("utf-8"));

													worksheet2.write_string(rowmap, col + 9, 'Direct');
													worksheet2.write_string(rowmap, col + 10, 'Phase 2');
													rowmap = rowmap + 1 

											if (objectType =='ABS Dev') :
												if row.cells[1].text.encode("utf-8") != row.cells[3].text.encode("utf-8") :
													Print("Insert ABS Dev Object: ", rownum_XSD, "XSD Name: ",xsdName , "Entity Name ", entityName, "Entity Type: Attribute")
													worksheet_XSD.write_string(rownum_XSD ,col + 0 , objectType.strip());
													worksheet_XSD.write_string(rownum_XSD ,col + 1 , 'XSD');

													# XSD Name from the input file
													worksheet_XSD.write_string(rownum_XSD ,col + 2 , xsdName.strip() );

													# XSD (target) attribute name from the word document table. process the string to extract the desired substring
													#worksheet_XSD.write_string(rownum_XSD, col + 3, entityName.strip())
													xsd_target_attribute = (row.cells[5].text.encode("utf-8")).strip()
													worksheet_XSD.write_string(rownum_XSD, col + 3, processXSDAttribute(xsd_target_attribute));

													worksheet_XSD.write_string(rownum_XSD ,col + 6 , 'Attribute');
													worksheet_XSD.write_string(rownum_XSD ,col + 7 , (row.cells[1].text.encode("utf-8")).strip());
													# nothing in col + 8 (URL)
													worksheet_XSD.write_string(rownum_XSD, col + 9,  removeSpaces((row.cells[5].text.encode("utf-8").strip())));
													worksheet_XSD.write_string(rownum_XSD, col + 10, 'Phase 2');
													rownum_XSD = rownum_XSD + 1

									#count is less than 0
									else :
										Print("Inside Else. Count: ", count)
										Print("Count is <=0. The header row in the IFS table selected")

										#insert the Entity Information. Entity type is either Interface or Class

										if(objectType =='IFS'):
											worksheet_IFS.write_string(rownum_IFS ,col + 0, objectType.strip());
											worksheet_IFS.write_string(rownum_IFS ,col + 1 , 'Default');
											worksheet_IFS.write_string(rownum_IFS, col + 2, ifsName.strip());
											worksheet_IFS.write_string(rownum_IFS, col + 3, "");
											worksheet_IFS.write_string(rownum_IFS ,col + 6 , 'Interface');
											worksheet_IFS.write_string(rownum_IFS ,col + 7 ,ifsDescription  );
											worksheet_IFS.write_string(rownum_IFS ,col + 8 , sharepoint_pathUrl );
											worksheet_IFS.write_string(rownum_IFS, col + 9, avaloq_service_name);
											worksheet_IFS.write_string(rownum_IFS, col + 10, 'Phase 2');
											rownum_IFS = rownum_IFS + 1

										if(objectType =='ABS Dev'):

											worksheet_XSD.write_string(rownum_XSD ,col + 0, objectType.strip())
											worksheet_XSD.write_string(rownum_XSD ,col + 1 , 'XSD')
											worksheet_XSD.write_string(rownum_XSD ,col + 2 ,xsdName.strip() )
											worksheet_XSD.write_string(rownum_XSD ,col + 3 , "")
											worksheet_XSD.write_string(rownum_XSD, col + 4, "")
											worksheet_XSD.write_string(rownum_XSD, col + 5, "")
											worksheet_XSD.write_string(rownum_XSD ,col + 6 , 'Class')
											worksheet_XSD.write_string(rownum_XSD, col + 7, ifsDescription)
											worksheet_XSD.write_string(rownum_XSD, col + 8, avaloq_service_name)
											worksheet_XSD.write_string(rownum_XSD, col + 9, "")
											worksheet_XSD.write_string(rownum_XSD, col + 10, 'Phase 2')
											rownum_XSD = rownum_XSD + 1

									count = count+ 1
 	 						ispara='false'


						if (isinstance(block, Paragraph)  ):
							Print("Paragraph found")

							if (sectionName != 'None') :
								#print block.text.encode("utf-8").strip()
								block_sectionName = block.text.encode("utf-8").strip()

								#print block_sectionName.upper()
								#print (sectionName.strip()).upper()

								if (block_sectionName.upper() == (sectionName.strip()).upper()):
									Print("Matching section found")
									ispara ='true'


# end of object_create() function


#create mapping excel sheet
outputDirectory = parentDir+'\\Output'
if not os.path.exists(outputDirectory):
    os.makedirs(outputDirectory)

fileName= outputDirectory + '\\mappingsheet' +  datetime.datetime.now().strftime("%Y%m%d-%H%M%S") +".xlsx"
workbook = xlsxwriter.Workbook(fileName)

worksheet_IFS = workbook.add_worksheet("IFS Objects")
worksheet_XSD = workbook.add_worksheet("XSD Objects")
worksheet2 = workbook.add_worksheet("Mapping Specification")
#sets the column width for the given range. E.g. set the columns from 0 to 8 a width of 30 characters

worksheet_IFS.set_column(0, 1 ,15)
worksheet_IFS.set_column(2, 3 ,30)
worksheet_IFS.set_column(4, 6 ,10)

worksheet_XSD.set_column(0, 1 ,15)
worksheet_XSD.set_column(2, 3 ,30)
worksheet_XSD.set_column(4, 6 ,10)

worksheet2.set_column(0, 1 ,15)
worksheet2.set_column(2, 3 ,30)
worksheet2.set_column(4, 6 ,10)



bold = workbook.add_format({'bold': 1});

worksheet_IFS.write('A1', 'System Name', bold);
worksheet_IFS.write('B1', 'Instance Name', bold)
worksheet_IFS.write('C1', 'Entity Name', bold)
worksheet_IFS.write('D1', 'Attribute Name', bold)
worksheet_IFS.write('E1', 'Owner', bold)
worksheet_IFS.write('F1', 'Parent', bold)
worksheet_IFS.write('G1', 'Type', bold)
worksheet_IFS.write('H1', 'Description', bold)
worksheet_IFS.write('I1', 'URL', bold)
worksheet_IFS.write('J1', 'XPATH', bold)
worksheet_IFS.write('K1', 'MDR Phase', bold)

worksheet_XSD.write('A1', 'System Name', bold);
worksheet_XSD.write('B1', 'Instance Name', bold)
worksheet_XSD.write('C1', 'Entity Name', bold)
worksheet_XSD.write('D1', 'Attribute Name', bold)
worksheet_XSD.write('E1', 'Owner', bold)
worksheet_XSD.write('F1', 'Parent', bold)
worksheet_XSD.write('G1', 'Type', bold)
worksheet_XSD.write('H1', 'Description', bold)
worksheet_XSD.write('I1', 'URL', bold)
worksheet_XSD.write('J1', 'XPATH', bold)
worksheet_XSD.write('K1', 'MDR Phase', bold)

worksheet2.write('A1', 'System Name', bold);
worksheet2.write('B1', 'Instance Name', bold)
worksheet2.write('C1', 'Entity Name', bold)
worksheet2.write('D1', 'Attribute Name', bold)
worksheet2.write('E1', 'System Name', bold);
worksheet2.write('F1', 'Instance Name', bold)
worksheet2.write('G1', 'Entity Name', bold)
worksheet2.write('H1', 'Attribute Name', bold)
worksheet2.write('I1', 'Description', bold)
worksheet2.write('J1', 'Business Rule', bold)
worksheet2.write('K1', 'MDR Phase', bold)



start_time = datetime.datetime.now()

print "Start Time:", start_time

print "Collecting Data from the input sheet..."

warnings.simplefilter("ignore")

#read excel sheet for input
try:

	wb = openpyxl.load_workbook(input_excel_filepath)#open input excel sheet
except IOError:
	print "ERROR: Could not open the input excel spreadsheet."
	print "ERROR DETAILS: Input file '" + input_excel_filepath +"' is missing. Please check the file name and directory details."
else:
	#open sheet by name
	sheet = wb.get_sheet_by_name('Sheet1')


	#print "No of rows in the input sheet:",sheet.max_row
	print "Harvesting objects and mappings from IFS documents. Please wait... "



	#IFS creates objects and mapping
	object_create('IFS')

	#ABS Dev creates objects only
	object_create('ABS Dev')

	end_time = datetime.datetime.now()


	print "End Time:  ",end_time
	print "Time taken to process the IFS documents:", end_time - start_time


	print "Task completed check your folder to latest Meta data mapping sheet file " + fileName

workbook.close()
