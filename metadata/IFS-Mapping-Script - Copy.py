
import docx
import openpyxl #install openpyxl first 
import os
import xlsxwriter 
import sys  
import datetime
import re
reload(sys)  
sys.setdefaultencoding('utf8')
from docx.document import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph
#parentDir = "..\\Desktop\\Metadata Repository" #change path of parent directory
parentDir = "C:\IFS"
rownum = 1
rowmap=1
col = 0
addtext = ""
def iter_block_items(parent):
    """
    Yield each paragraph and table child within *parent*, in document order.
    Each returned value is an instance of either Table or Paragraph. *parent*
    would most commonly be a reference to a main Document object, but
    also works for a _Cell object, which itself can contain paragraphs and tables.
    """
    if isinstance(parent, Document):
        parent_elm = parent.element.body
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        raise ValueError("something's not right")

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)
def object_create(objectType):
	for row in range(2, sheet.max_row):
		if ( sheet['C' + str(row)].value != None):
			if(sheet['B' + str(row)].value == "Open"):
				#get name of ifs doc to read 
				global parentDir
				docName =  parentDir+"\\"+ sheet['C' + str(row)].value 
				#read table form ifs doc 
				doc1 = docx.Document(docName)
				
				sectionName = sheet['E' + str(row)].value.encode("utf-8").strip()
				ifsName=sheet['D' + str(row)].value.encode("utf-8").strip()
				xsdName=sheet['I' + str(row)].value.encode("utf-8").strip()
				ifsDescription = sheet['F' + str(row)].value.encode("utf-8").strip()
				pathUrl= sheet['G' + str(row)].value
				#object creation for ifs 
				ispara = 'false'
				global rownum
				global rowmap
				for block in iter_block_items(doc1):
						if (isinstance(block, Table)):
							if ispara == 'true' :
								ifsTable= block
								count = 0
								entityList=[];
								for row in ifsTable.rows:
									entityName= re.sub("[\(\[].*?[\)\]]", "", row.cells[0].text.encode("utf-8"))
									if (count>0):
										if(row.cells[2].text.encode("utf-8") != "Aggregate") :
											if (objectType== 'IFS'):
												if row.cells[1].text.encode("utf-8") != row.cells[3].text.encode("utf-8") :
													if row.cells[0].text.encode("utf-8") not in entityList :
													
														worksheet.write_string(rownum ,col , objectType);
														worksheet.write_string(rownum ,col + 1 , 'Default');
														worksheet.write_string(rownum ,col + 2 ,ifsName );
														entityList.append(row.cells[0].text.encode("utf-8"))
														worksheet.write_string(rownum ,col + 3 , entityName)
														worksheet.write_string(rownum ,col + 6 , 'Attribute');
														worksheet.write_string(rownum ,col + 7 ,  row.cells[1].text.encode("utf-8"));
														rownum = rownum + 1
													worksheet2.write_string(rowmap ,col , 'IFS');
													worksheet2.write_string(rowmap ,col+1 , 'Default');
													worksheet2.write_string(rowmap ,col+2 , ifsName);
													worksheet2.write_string(rowmap ,col+3 ,  entityName);
													worksheet2.write_string(rowmap ,col +4 , 'ABS Dev');
													worksheet2.write_string(rowmap ,col + 5 , 'XSD');
													worksheet2.write_string(rowmap ,col + 6 , xsdName);
													worksheet2.write_string(rowmap ,col + 7 ,entityName);
													worksheet2.write_string(rowmap ,col + 8 ,row.cells[1].text.encode("utf-8"));
													rowmap = rowmap + 1	
											if (objectType =='ABS Dev') :
												if row.cells[1].text.encode("utf-8") != row.cells[3].text.encode("utf-8") :
													worksheet.write_string(rownum ,col , objectType);
													worksheet.write_string(rownum ,col + 1 , 'XSD');
													worksheet.write_string(rownum ,col + 2 ,xsdName );
													worksheet.write_string(rownum ,col + 9 ,  row.cells[5].text.encode("utf-8"));
													worksheet.write_string(rownum ,col + 3 , entityName)
													worksheet.write_string(rownum ,col + 6 , 'Attribute');
													worksheet.write_string(rownum ,col + 7 ,  row.cells[1].text.encode("utf-8"));
													rownum = rownum + 1
											
									else :
										
										if(objectType =='IFS') :
											worksheet.write_string(rownum ,col , objectType);
											worksheet.write_string(rownum ,col + 1 , 'Default');
											worksheet.write_string(rownum ,col + 6 , 'Interface');
											worksheet.write_string(rownum ,col + 2 ,ifsName );
											worksheet.write_string(rownum ,col + 7 ,ifsDescription  );
											worksheet.write_string(rownum ,col + 8 , pathUrl )
											worksheet.write_string(rownum ,col + 3 , "");
										if(objectType =='ABS Dev') :
											worksheet.write_string(rownum ,col , objectType);
											worksheet.write_string(rownum ,col + 1 , 'XSD');
											worksheet.write_string(rownum ,col + 6 , 'Class');
											worksheet.write_string(rownum ,col + 2 ,xsdName );
											worksheet.write_string(rownum ,col + 3 , "");
										
										rownum = rownum + 1
									count = count+ 1
							ispara='false'
						if (isinstance(block, Paragraph)  ):
							if (sectionName != 'None') :
								if (block.text.encode("utf-8").strip() == sectionName):
										ispara ='true'
							
#create mapping excel sheet 
outputDirectory = parentDir+'\\Output'
if not os.path.exists(outputDirectory):
    os.makedirs(outputDirectory)
fileName= outputDirectory + '\\mappingsheet' +  datetime.datetime.now().strftime("%Y%m%d-%H%M%S") +".xlsx"
workbook = xlsxwriter.Workbook(fileName)
worksheet = workbook.add_worksheet("Object Definition")
worksheet2 = workbook.add_worksheet("Mapping Specification")
worksheet2.set_column(0, 8 ,30)
bold = workbook.add_format({'bold': 1});
worksheet.write('A1', 'System Name', bold);
worksheet.write('B1', 'Instance Name', bold)
worksheet.write('C1', 'Entity Name', bold)
worksheet.write('D1', 'Attribute Name', bold)
worksheet.write('E1', 'Owner', bold)
worksheet.write('F1', 'Parent', bold)
worksheet.write('G1', 'Type', bold)
worksheet.write('H1', 'Description', bold)
worksheet.write('I1', 'URL', bold)
worksheet.write('J1', 'XPATH', bold)
worksheet2.write('A1', 'System Name', bold);
worksheet2.write('B1', 'Instance Name', bold)
worksheet2.write('C1', 'Entity Name', bold)
worksheet2.write('D1', 'Attribute Name', bold)
worksheet2.write('E1', 'System Name', bold);
worksheet2.write('F1', 'Instance Name', bold)
worksheet2.write('G1', 'Entity Name', bold)
worksheet2.write('H1', 'Attribute Name', bold)
worksheet2.write('I1', 'Description', bold)
worksheet.set_column(0, 8 ,30)
print "Collecting Data from IFS "
#read excel sheet for input 
wb = openpyxl.load_workbook(parentDir+'\\IFS Documents Analysis.xlsx')#open input excel sheet
sheet = wb.get_sheet_by_name('Sheet1')
#open sheet by name 
print "Creating object sheet and mapping sheet in excel... "
object_create('IFS')
object_create('ABS Dev')
print "Task completed check your folder to latest Meta data mapping sheet file " + fileName
workbook.close()	